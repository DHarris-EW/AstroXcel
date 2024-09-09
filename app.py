from copy import copy
import sv_ttk
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class App(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        container = tk.Frame(self)
        container.grid(padx=20, pady=5)

        # Variables for merging
        self.mergeFileNames = tk.StringVar()
        self.columnName = tk.StringVar()
        self.mergeFilesPath = []
        self.merge_ws = []
        
        # Variables for output directory
        self.outputDirName = tk.StringVar()
        self.outPutDirPath = ""

        # Variables for the original
        self.originalFileName = tk.StringVar()
        self.originalWS = {}
        self.originalWB = {}

        # Variables for orignial file when column <Count_ID> is added
        self.countIDFileName = tk.StringVar()
        self.countIDWS = {}
        self.countIDWB = {}

        # Different frames
        CreateFiles(container, self).grid(row=1, column=0, pady=5, padx=5, sticky="nesw")
        MergeFiles(container, self).grid(row=0, column=1, rowspan=2, pady=5, padx=5, sticky="nesw")
        Options(container, self).grid(row=0, column=0, pady=5, padx=5, sticky="nesw")

    def SelectDirectory(self, event=None):
        # Selecting the output directory
        self.outPutDirPath = str(filedialog.askdirectory())
        self.outputDirName.set(self.outPutDirPath)

    def UploadAction(self, event=None, countIDFile=False):
        # upladoing the orignial Excel file or orinigal file with <Count_ID> column
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

        # Loading the workbook and setting the variable according the Excel file uploaded
        if not countIDFile:
            self.originalFileName.set(file_path.split("/")[-1])
            self.originalWB = load_workbook(file_path, data_only=True)
            self.originalWS = self.originalWB.active
        else:
            self.countIDFileName.set(file_path.split("/")[-1])
            self.countIDWB = load_workbook(file_path, data_only=True)
            self.countIDWS = self.countIDWB.active
        

    def UploadActionMultiple(self, event=None):
        # Upload multiple files that are selecting when mergine
        files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])

        # Adds all the file paths to a variable
        for path in files:
            file_paths += path.split("/")[-1] + "\n"
            self.mergeFilesPath.append(path)

        self.mergeFileNames.set(file_paths)

    def CreateDataFrame(self, type):
        # Creates a dataframe so data can be grouped and column names can be read quickly
        if type=="countID":
            data = self.countIDWS.values
            columns = next(data)
            return pd.DataFrame(data, columns=columns)
        elif type=="original":
            data = self.originalWS.values
            columns = next(data)
            return pd.DataFrame(data, columns=columns)


    def CreateWorkBooks(self):
        df = self.CreateDataFrame(type="original")
        if self.outPutDirPath:
            # Create a 'copy' of the original file with the column 'Count_ID'
            # Count_ID used to keep track of row loction when splitting and merging
            if "Count_ID" not in df.columns:
                df.insert(0, "Count_ID", range(1, len(df) + 1))
                df.to_excel(self.outPutDirPath + "/" + "CountID_" + self.originalFileName.get(), index=False)
            # ColumnName is used to determine what column to group the data on
            if self.columnName.get() not in df.columns:
                if "Zuordnung" in df.columns:
                    self.columnName.set("Zuordnung")
                else:
                    self.columnName.set = df.columns[0]
            # Grouping data based on column name
            for groupValue, groupDF in df.groupby(self.columnName.get()):
                file_name = self.outPutDirPath + "/" + str(groupValue) + ".xlsx"
                groupDF.to_excel(file_name, index=False)
                wb = load_workbook(file_name)
                ws = wb.active
                # Stlying the workbook
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    if col[0].value == "Count_ID":
                        ws.column_dimensions[get_column_letter(col[0].column)].hidden= True
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
                wb.save(file_name)
        else:
            messagebox.showerror("Missing Output file", "No Output file selected.")

    def MergeWorkBooks(self):
        if self.mergeFilesPath:
            for path in self.mergeFilesPath:
                wb = load_workbook(path)
                ws = wb.active
                self.merge_ws.append(ws)
            
            df = self.CreateDataFrame(type="countID")

            if self.countIDWB and "Count_ID" in df.columns:
                if self.outPutDirPath:
                    update_cells = {}

                    for ws_idx, ws in enumerate(self.merge_ws):
                        for row_idx, row in enumerate(ws.rows):
                            update_cells[row[0].value] = {"ws_idx": ws_idx, "row_idx": row_idx + 1}


                    for col in self.merge_ws[0].columns:
                        max_length = max(len(str(cell.value)) for cell in col)

                        self.countIDWS.column_dimensions[get_column_letter(col[0].column)].width = max_length

                    for row in self.countIDWS.rows:
                        if row[0].value in update_cells:
                            for cell_new, cell in zip(self.merge_ws[update_cells[row[0].value]["ws_idx"]][update_cells[row[0].value]["row_idx"]], row):
                                    if cell_new.has_style:
                                        cell.font = copy(cell_new.font)
                                        cell.border = copy(cell_new.border)
                                        cell.fill = copy(cell_new.fill)
                                        cell.number_format = copy(cell_new.number_format)
                                        cell.protection = copy(cell_new.protection)
                                        cell.alignment = copy(cell_new.alignment)
                                    cell.value = cell_new.value

                    self.countIDWS.delete_cols(1)
                    self.countIDWB.save(self.outPutDirPath + "/" + "AstroXCel.xlsx")
                else:
                    messagebox.showerror("Missing Output file", "No Output file selected.")
            else:
                messagebox.showerror("Missing Master file", "No Master file selected. <Coud_ID> column must be included.")
        else:
            messagebox.showerror("Missing Merge files", "No files to merge selected.")

class Options(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, highlightbackground="gray24", highlightthickness=2, padx=5, pady=5)
        self.columnconfigure(0, weight=1)

        ttk.Label(self, text="Output Folder Name", anchor="center").grid(row=1, column=0, padx=5, pady=10, sticky="nesw")
        ttk.Label(self, textvariable=controller.outputDirName, anchor="center").grid(row=2, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Button(self, text="Select Folder", command=controller.SelectDirectory).grid(row=3, column=0, padx=5, pady=5, sticky="nswe")

class CreateFiles(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, highlightbackground="gray24", highlightthickness=2, padx=5, pady=5)
        self.columnconfigure(0, weight=1)

        ttk.Label(self, text="File Name", anchor="center").grid(row=0, column=0, padx=5, pady=10, sticky="nesw")
        ttk.Label(self, textvariable=controller.originalFileName, anchor="center").grid(row=1, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Button(self, text="Select File", command=controller.UploadAction).grid(row=2, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Label(self, text="Column Name", anchor="center").grid(row=3, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Entry(self, textvariable=controller.columnName).grid(row=4, column=0, padx=5,pady=5, sticky="nesw")
        ttk.Button(self, text="Create Groups", command=controller.CreateWorkBooks).grid(row=5, column=0, padx=5, pady=5, sticky="nesw")

class MergeFiles(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, highlightbackground="gray24", highlightthickness=2, padx=5, pady=5)
        self.columnconfigure(0, weight=1)

        ttk.Label(self, text="CountID File Name", anchor="center").grid(row=0, column=0, padx=5, pady=10, sticky="nesw")
        ttk.Label(self, textvariable=controller.countIDFileName, anchor="center").grid(row=1, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Button(self, text="Select File", command= lambda: controller.UploadAction(countIDFile=True)).grid(row=2, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Label(self, text="File Names", anchor="center").grid(row=3, column=0, padx=5, pady=10, sticky="nesw")
        ttk.Label(self, textvariable=controller.mergeFileNames, anchor="center").grid(row=4, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Button(self, text="Select Merge Files", command=controller.UploadActionMultiple).grid(row=5, column=0, padx=5, pady=5, sticky="nesw")
        ttk.Button(self, text="Merge", command=controller.MergeWorkBooks).grid(row=6, column=0, padx=5, pady=5, sticky="nesw")

app = App()
sv_ttk.set_theme("dark")
app.title("AstroXCel")
app.resizable(False, False)
app.mainloop()